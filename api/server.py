import datetime

import openpyxl as openpyxl
import xlrd
import xlwt
from openpyxl import load_workbook, Workbook
from xlutils3 import copy

from model.model import *


def query_city_list():
    query = db.session().query(CityName).all()
    return json.dumps([i.to_json() for i in query], ensure_ascii=False)


def query_month_data_by_city(city):
    query = db.session().query(MonthData).filter_by(cityName=city).all()
    return json.dumps([i.to_json() for i in query], ensure_ascii=False)


# query_day_data_by_city

def query_day_data_by_city(city, month):
    query = db.session().query(DayData).filter(DayData.time_point.ilike(month + '%')).filter_by(cityName=city).all()
    return json.dumps([i.to_json() for i in query], ensure_ascii=False)


# query_all_day_data
def query_all_day_data(city):
    query = db.session().query(DayData).filter_by(cityName=city).all()
    return json.dumps([i.to_json() for i in query], ensure_ascii=False)


def query_lately_aqi_city(provinces_id):
    today = datetime.date.today()
    oneday = datetime.timedelta(days=1)
    yesterday = today - oneday
    now_y, now_m, now_d = yesterday.strftime('%Y-%m-%d').split('-')
    s = now_y + '-' + now_m + '-' + now_d
    result = {}
    if provinces_id is not None:
        query = db.session().query(DayData, CityName).filter(DayData.time_point == s).filter(
            DayData.cityName == CityName.city).filter(CityName.regionid.ilike(provinces_id[:2] + '%')).all()
    else:
        query = db.session().query(DayData, CityName).filter(DayData.time_point == s).filter(
            DayData.cityName == CityName.city).all()

    for q in query:
        result[q[1].regionid] = [q[0].cityName, q[0].aqi]

    if provinces_id is None or provinces_id == '650000':
        result[652800] = result[652900]

    return json.dumps(result, ensure_ascii=False)


# 写文件测试
# def write_excel(result, ci):
#     wbk = xlwt.Workbook()
#     for r in result:
#         sheet = wbk.add_sheet(r)
#         target_type = target_type = ['time_point', 'aqi', 'pm2_5', 'pm10', 'so2', 'no2', 'co', 'o3', 'rank', 'quality']
#         for k, v in enumerate(target_type):
#             sheet.write(0, k, v)
#         for k, v in enumerate(result[r]):
#             sheet.write(k + 1, 0, v['time_point'])
#             sheet.write(k + 1, 1, v['aqi'])
#             sheet.write(k + 1, 2, v['pm2_5'])
#             sheet.write(k + 1, 3, v['pm10'])
#             sheet.write(k + 1, 4, v['so2'])
#             sheet.write(k + 1, 5, v['no2'])
#             sheet.write(k + 1, 6, v['co'])
#             sheet.write(k + 1, 7, v['o3'])
#             sheet.write(k + 1, 8, v['rank'])
#             sheet.write(k + 1, 9, v['quality'])
#     wbk.save('static/temp/' + ci + '.xls')
#     return 'static/temp/' + ci + '.xls'
#
#
# def query_provinces_aqi(provinces_id):
#     query = db.session().query(DayData, CityName).filter(DayData.cityName == CityName.city).filter(
#         CityName.regionid.ilike(provinces_id[:2] + '%')).all()
#     temp = [i[0].to_json() for i in query]
#     result = {}
#     for t in temp:
#         if t['cityName'] not in result:
#             result[t['cityName']] = []
#         result[t['cityName']].append(t)
#     return write_excel(result, provinces_id)
#
#
# def query_year_data(year):
#     query = db.session().query(DayData).filter(DayData.time_point.contains(year + '%')).all()
#     temp = [i.to_json() for i in query]
#     result = {}
#     for t in temp:
#         if t['cityName'] not in result:
#             result[t['cityName']] = []
#         result[t['cityName']].append(t)
#     return write_excel(result, year)
def query_day_data(day):
    query = db.session().query(DayData).filter(DayData.time_point == day).all()
    temp = [i.to_json() for i in query]

    return json.dumps(temp, ensure_ascii=False)


def update_year_excel(temp):
    wb = load_workbook('static/temp/2019.xlsx')
    for t in temp:
        sheet = wb.get_sheet_by_name(t['cityName'])
        # print(sheet2)#根据表单名字，获取Excel表内容。返回的是一个存放地址
        sheet.append([
            t['time_point'],
            t['aqi'],
            t['pm2_5'],
            t['pm10'],
            t['so2'],
            t['no2'],
            t['co'],
            t['o3'],
            t['rank'],
            t['quality'],
        ])
    print('更新2019')
    wb.save('static/temp/2019.xlsx')


def update_provinces_excel(temp, ci):
    wb = load_workbook('static/temp/' + ci + '0000' + '.xlsx')
    for i in temp:
        t = temp[i][0]
        sheet = wb.get_sheet_by_name(t['cityName'])
        sheet.append([
            t['time_point'],
            t['aqi'],
            t['pm2_5'],
            t['pm10'],
            t['so2'],
            t['no2'],
            t['co'],
            t['o3'],
            t['rank'],
            t['quality'],
        ])
    print('更新' + ci)
    wb.save('static/temp/' + ci + '0000' + '.xlsx')


def create_excel(result, year):
    wb = Workbook()
    for r in result:
        wb.create_sheet(r)
        sheet = wb.get_sheet_by_name(r)
        target_type = ['time_point', 'aqi', 'pm2_5', 'pm10', 'so2', 'no2', 'co', 'o3', 'rank', 'quality']
        sheet.append(target_type)
        for t in result[r]:
            sheet.append([
                t['time_point'],
                t['aqi'],
                t['pm2_5'],
                t['pm10'],
                t['so2'],
                t['no2'],
                t['co'],
                t['o3'],
                t['rank'],
                t['quality'],
            ])
    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
    wb.save('static/temp/' + year + '.xlsx')


def query_update_day_excel():
    today = datetime.date.today()
    oneday = datetime.timedelta(days=1)
    yesterday = today - oneday
    now_date = yesterday.strftime('%Y-%m-%d')
    proid = [11, 12, 13, 14, 15, 21, 22, 23, 31, 32, 33, 34, 35, 36, 37, 41, 42, 43, 44, 45, 46, 50, 51, 52, 53, 54, 61,
             62, 63, 64, 65]
    query = db.session().query(DayData).filter(DayData.time_point == now_date).all()
    temp = [i.to_json() for i in query]
    update_year_excel(temp)
    for y in proid:
        query = db.session().query(DayData, CityName).filter(DayData.cityName == CityName.city).filter(
            CityName.regionid.ilike(str(y) + '%')).filter(DayData.time_point == now_date).all()
        temp = [i[0].to_json() for i in query]
        result = {}
        for t in temp:
            if t['cityName'] not in result:
                result[t['cityName']] = []
            result[t['cityName']].append(t)
        update_provinces_excel(result, str(y))
    return 'sucess'
